"""
reporter.py
===========
Generates text report (for Studio validation) and Excel export.
All heavy imports INSIDE functions to avoid sandbox restrictions.
"""

import mistralai.workflows as wfk

LEVEL_EMOJI = {
    "Critical": "🔴",
    "High":     "🟠",
    "Medium":   "🟡",
    "Low":      "🟢",
}

SCORE_LABEL = {
    1: "Non-compliant",
    2: "Partially compliant",
    3: "Compliant",
    4: "Mature",
}

DECISION_EMOJI = {
    "Approved":                 "✅",
    "Approved with conditions": "⚠️",
    "Rejected":                 "❌",
}


@wfk.activity()
async def generate_text_report(
    analysis: dict,
    scoring: dict,
    vendor: str,
    project: str,
    analyst: str,
) -> str:
    """Generates a plain text summary shown in Studio for human validation."""
    from datetime import datetime
    from workflows.analyzer import RiskAnalysis
    from workflows.scorer import ScoringResult

    analysis = RiskAnalysis(**analysis) if isinstance(analysis, dict) else analysis

    if isinstance(scoring, dict):
        scoring = ScoringResult(**{k: v for k, v in scoring.items() if k in ScoringResult.model_fields})

    decision_emoji = DECISION_EMOJI.get(analysis.final_decision, "❓")
    overall_emoji  = LEVEL_EMOJI.get(analysis.overall_score, "⚪")

    lines = [
        f"=== TPRA REPORT — {vendor} ===",
        f"Project       : {project}",
        f"Date          : {datetime.now().strftime('%Y-%m-%d')}",
        f"Risk Level    : {overall_emoji} {analysis.overall_score}",
        f"Global Score  : {scoring.global_score:.1f} / 4.0",
        f"Final Decision: {decision_emoji} {analysis.final_decision}",
        f"Showstoppers  : {scoring.showstopper_count}",
        "",
        "--- EXECUTIVE SUMMARY ---",
        analysis.executive_summary,
        "",
    ]

    if analysis.showstoppers:
        lines += ["--- 🚨 SHOWSTOPPERS ---"]
        for s in analysis.showstoppers:
            lines.append(f"  ❌ {s}")
        lines.append("")

    lines += [f"--- IDENTIFIED RISKS ({len(analysis.risks)}) ---"]
    for i, risk in enumerate(analysis.risks, 1):
        e = LEVEL_EMOJI.get(risk.level, "⚪")
        ids = ", ".join(risk.related_question_ids) if risk.related_question_ids else "—"
        lines += [
            f"{i}. [{e} {risk.level}] {risk.title}",
            f"   📎 Questions: {ids}",
            f"   → {risk.description}",
            f"   ✅ {risk.recommendation}",
            "",
        ]

    lines += [f"--- QUESTION SCORES ({len(scoring.question_scores)} questions) ---"]
    for q in scoring.question_scores:
        label = SCORE_LABEL.get(q.score, "?")
        ss    = " 🚨 SHOWSTOPPER" if q.is_showstopper else ""
        lines += [f"[{q.question_id}] {q.score}/4 — {label}{ss}"]
        if q.follow_up_question:
            lines.append(f"   ❓ {q.follow_up_question}")
        lines.append("")

    lines.append(f"Prepared by: {analyst}")
    return "\n".join(lines)


@wfk.activity()
async def export_excel(
    analysis: dict,
    scoring: dict,
    vendor: str,
    project: str,
    analyst: str,
) -> str:
    """
    Generates the final Excel report with multiple sheets:
    - Summary         : key metrics and decision
    - Question Scores : per-question scores, flags, follow-ups, linked risks
    - Risks           : all identified risks with linked question IDs
    - Showstoppers    : critical blocking issues
    """
    import os
    import tempfile
    from datetime import datetime
    from workflows.analyzer import RiskAnalysis
    from workflows.scorer import ScoringResult
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    analysis = RiskAnalysis(**analysis) if isinstance(analysis, dict) else analysis

    if isinstance(scoring, dict):
        scoring = ScoringResult(**{k: v for k, v in scoring.items() if k in ScoringResult.model_fields})

    # ── Color palette ────────────────────────────────────────────────────────
    COLORS = {
        "Critical": "C0392B",
        "High":     "E67E22",
        "Medium":   "F1C40F",
        "Low":      "27AE60",
        "Approved": "27AE60",
        "Approved with conditions": "E67E22",
        "Rejected": "C0392B",
        1: "C0392B",
        2: "E67E22",
        3: "F1C40F",
        4: "27AE60",
        "header":    "1A252F",
        "subheader": "2C3E50",
        "light":     "ECF0F1",
        "white":     "FFFFFF",
    }

    def fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def bold_font(color="FFFFFF", size=11) -> Font:
        return Font(bold=True, color=color, size=size)

    def header_style(ws, row, cols, bg="1A252F"):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill(bg)
            cell.font = bold_font("FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def set_col_width(ws, col, width):
        ws.column_dimensions[get_column_letter(col)].width = width

    thin = Side(style="thin", color="BDC3C7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def apply_border(ws, min_row, max_row, min_col, max_col):
        for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                 min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = border

    # ── Build cross-reference: question_id → list of risk titles ────────────
    # Used to populate "Linked Risks" column in Question Scores sheet
    question_to_risks: dict[str, list[str]] = {}
    for risk in analysis.risks:
        for qid in risk.related_question_ids:
            question_to_risks.setdefault(qid, []).append(risk.title)

    # ── Create workbook ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Summary
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

    def add_kv(label, value, val_bg=None):
        r = ws.max_row + 1
        ws.cell(r, 1, label).font = bold_font("1A252F", 10)
        ws.cell(r, 1).fill = fill("ECF0F1")
        cell = ws.cell(r, 2, str(value))
        cell.alignment = Alignment(wrap_text=True)
        if val_bg:
            cell.fill = fill(val_bg)
            cell.font = bold_font("FFFFFF", 10)
        apply_border(ws, r, r, 1, 2)

    ws.merge_cells("A1:B1")
    ws["A1"] = "TPRA — Third Party Risk Assessment Report"
    ws["A1"].font = bold_font("FFFFFF", 14)
    ws["A1"].fill = fill("1A252F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.append([])  # spacer

    add_kv("Vendor",           vendor)
    add_kv("Project",          project)
    add_kv("Analyst",          analyst)
    add_kv("Date",             datetime.now().strftime("%Y-%m-%d %H:%M"))
    add_kv("Risk Level",       analysis.overall_score,
           COLORS.get(analysis.overall_score, "555555"))
    add_kv("Global Score",     f"{scoring.global_score:.1f} / 4.0",
           COLORS.get(round(scoring.global_score), "555555"))
    add_kv("Final Decision",   analysis.final_decision,
           COLORS.get(analysis.final_decision, "555555"))
    add_kv("Showstoppers",     scoring.showstopper_count,
           "C0392B" if scoring.showstopper_count > 0 else "27AE60")
    add_kv("Questions scored", len(scoring.question_scores))
    add_kv("Scores ≤ 2",       scoring.low_score_count,
           "C0392B" if scoring.low_score_count > 0 else "27AE60")

    ws.append([])
    r = ws.max_row + 1
    ws.merge_cells(f"A{r}:B{r}")
    ws.cell(r, 1, "Executive Summary").font = bold_font("FFFFFF", 11)
    ws.cell(r, 1).fill = fill("2C3E50")
    ws.cell(r, 1).alignment = Alignment(horizontal="center")

    r = ws.max_row + 1
    ws.merge_cells(f"A{r}:B{r}")
    ws.cell(r, 1, analysis.executive_summary)
    ws.cell(r, 1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 60

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 2 — Question Scores
    # ── Nouvelle colonne "Linked Risks" pour retrouver les risques par ID ───
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Question Scores")
    headers = [
        "ID",
        "Question",
        "Vendor Response",
        "Score",
        "Level",
        "Justification",
        "Linked Risks",        # ← NOUVEAU : risques liés à cette question
        "Showstopper",
        "Flag Reason",
        "Follow-up Question",
    ]
    widths = [12, 35, 40, 8, 18, 35, 45, 12, 30, 40]

    ws2.append(headers)
    header_style(ws2, 1, len(headers))
    ws2.row_dimensions[1].height = 20
    for i, w in enumerate(widths, 1):
        set_col_width(ws2, i, w)

    for q in scoring.question_scores:
        label = SCORE_LABEL.get(q.score, "?")

        # Risques liés à cette question (via le cross-ref construit plus haut)
        linked_risk_titles = question_to_risks.get(q.question_id, [])
        linked_risks_str = "\n".join(f"• {t}" for t in linked_risk_titles) if linked_risk_titles else "—"

        row = [
            q.question_id,
            q.question,
            q.response,
            q.score,
            label,
            q.justification,
            linked_risks_str,  # ← colonne 7
            "🚨 YES" if q.is_showstopper else "No",
            q.flag_reason,
            q.follow_up_question,
        ]
        ws2.append(row)
        r = ws2.max_row

        # Score column color (col 4)
        sc = ws2.cell(r, 4)
        sc.fill = fill(COLORS.get(q.score, "FFFFFF"))
        sc.font = bold_font("FFFFFF", 10)
        sc.alignment = Alignment(horizontal="center")

        # Showstopper highlight (col 8)
        if q.is_showstopper:
            ws2.cell(r, 1).fill = fill("FADBD8")
            ws2.cell(r, 8).fill = fill("C0392B")
            ws2.cell(r, 8).font = bold_font("FFFFFF", 10)

        # Wrap text for long columns
        for col in [2, 3, 6, 7, 9, 10]:
            ws2.cell(r, col).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[r].height = max(40, 15 * len(linked_risk_titles)) if linked_risk_titles else 40

    apply_border(ws2, 1, ws2.max_row, 1, len(headers))
    ws2.freeze_panes = "A2"

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 3 — Risks
    # ── Nouvelle colonne "Question IDs" pour retrouver les questions par risque
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Risks")
    risk_headers = ["#", "Title", "Level", "Question IDs", "Description", "Recommendation"]
    risk_widths  = [5, 30, 12, 30, 50, 50]

    ws3.append(risk_headers)
    header_style(ws3, 1, len(risk_headers))
    ws3.row_dimensions[1].height = 20
    for i, w in enumerate(risk_widths, 1):
        set_col_width(ws3, i, w)

    for i, risk in enumerate(analysis.risks, 1):
        # IDs des questions liées, séparés par des sauts de ligne
        q_ids_str = "\n".join(risk.related_question_ids) if risk.related_question_ids else "—"

        ws3.append([i, risk.title, risk.level, q_ids_str, risk.description, risk.recommendation])
        r = ws3.max_row

        # Level column color (col 3)
        lc = ws3.cell(r, 3)
        lc.fill = fill(COLORS.get(risk.level, "FFFFFF"))
        lc.font = bold_font("FFFFFF", 9)
        lc.alignment = Alignment(horizontal="center", vertical="top")

        # Question IDs column : centré et wrap
        qc = ws3.cell(r, 4)
        qc.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
        qc.font = Font(bold=True, color="2C3E50", size=9)

        # Description & Recommendation : wrap
        for col in [5, 6]:
            ws3.cell(r, col).alignment = Alignment(wrap_text=True, vertical="top")

        n_ids = len(risk.related_question_ids) if risk.related_question_ids else 1
        ws3.row_dimensions[r].height = max(45, 15 * n_ids)

    apply_border(ws3, 1, ws3.max_row, 1, len(risk_headers))
    ws3.freeze_panes = "A2"

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 4 — Showstoppers
    # ════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Showstoppers")
    ws4.merge_cells("A1:C1")
    ws4["A1"] = "🚨 Showstoppers — Critical Blocking Issues"
    ws4["A1"].font = bold_font("FFFFFF", 12)
    ws4["A1"].fill = fill("C0392B")
    ws4["A1"].alignment = Alignment(horizontal="center")
    ws4.row_dimensions[1].height = 25

    if analysis.showstoppers:
        ws4.append(["#", "Description", "Action Required"])
        header_style(ws4, 2, 3, "2C3E50")
        set_col_width(ws4, 1, 5)
        set_col_width(ws4, 2, 60)
        set_col_width(ws4, 3, 40)
        for i, ss in enumerate(analysis.showstoppers, 1):
            ws4.append([i, ss, "Immediate action required before proceeding."])
            r = ws4.max_row
            ws4.cell(r, 1).fill = fill("FADBD8")
            ws4.cell(r, 2).alignment = Alignment(wrap_text=True)
            ws4.cell(r, 3).alignment = Alignment(wrap_text=True)
            ws4.row_dimensions[r].height = 35
        apply_border(ws4, 2, ws4.max_row, 1, 3)

    ss_questions = [q for q in scoring.question_scores if q.is_showstopper]
    if ss_questions:
        ws4.append([])
        r = ws4.max_row + 1
        ws4.merge_cells(f"A{r}:C{r}")
        ws4.cell(r, 1, "Questions Flagged as Showstoppers")
        ws4.cell(r, 1).font = bold_font("FFFFFF", 11)
        ws4.cell(r, 1).fill = fill("2C3E50")
        ws4.cell(r, 1).alignment = Alignment(horizontal="center")

        ws4.append(["ID", "Question", "Score / Follow-up"])
        header_style(ws4, ws4.max_row, 3, "555555")
        for q in ss_questions:
            fu = f" | ❓ {q.follow_up_question}" if q.follow_up_question else ""
            ws4.append([q.question_id, q.question, f"{q.score}/4 — {SCORE_LABEL.get(q.score, '?')}{fu}"])
            r = ws4.max_row
            ws4.cell(r, 1).fill = fill("FADBD8")
            for col in [2, 3]:
                ws4.cell(r, col).alignment = Alignment(wrap_text=True)
            ws4.row_dimensions[r].height = 40

    # ── Save ─────────────────────────────────────────────────────────────────
    filename = f"TPRA_{vendor.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb.save(filepath)
    return filepath